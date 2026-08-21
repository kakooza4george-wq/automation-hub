import os, time, re, glob
import streamlit as st
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from io import BytesIO
import difflib

st.set_page_config(page_title="Automation Hub", layout="wide", page_icon="🤖")

st.markdown("""
<style>
    .stApp { background-color: #F3F3F3; color: #000000; font-family: 'Segoe UI', sans-serif; }
    h1, h2, h3 { color: #0078D4; font-weight: 600; }
    .stButton > button {
        background-color: #0078D4; color: white; border-radius: 4px;
        border: none; padding: 8px 16px; font-weight: 500;
    }
    .stButton > button:hover { background-color: #106EBE; }
    p, li, span, div { color: #000000 !important; }
    .log-box {
        background: #1e1e1e; color: #00ff88; font-family: monospace;
        font-size: 13px; padding: 12px 16px; border-radius: 6px;
        max-height: 400px; overflow-y: auto; white-space: pre-wrap;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 1: Excel Stock Movement Filler
# ─────────────────────────────────────────────────────────────────────────────

def normalize_name(name):
    if not name:
        return ""
    return str(name).upper().replace(" ", "").replace("S", "")

def excel_serial_to_date(serial):
    if isinstance(serial, (float, int)):
        return (datetime(1899, 12, 30) + timedelta(days=serial)).date()
    if isinstance(serial, datetime):
        return serial.date()
    return None

def process_excel(template_file, report_file, damages_file, output_name="filled_template.xlsx"):
    try:
        template_bytes = template_file.read()
        wb_temp = openpyxl.load_workbook(BytesIO(template_bytes))
        ws_temp = wb_temp['Sheet1']
        template_first_date = None
        for r in range(1, ws_temp.max_row + 1):
            date_val = ws_temp.cell(r, 1).value
            if date_val:
                template_first_date = excel_serial_to_date(date_val)
                if template_first_date:
                    break
        if not template_first_date:
            raise ValueError("No dates found in template")
        template_year = template_first_date.year
        report_df  = pd.read_excel(report_file)
        damages_df = pd.read_excel(damages_file)
        report_df['date'] = pd.to_datetime(report_df['date'], dayfirst=True)
        report_df['date'] = report_df['date'].apply(lambda d: d.replace(year=template_year))
        report_df_sorted = report_df.sort_values(['abbreviations', 'date'])
        adj_df = report_df_sorted[report_df_sorted['movement_type'] == 'Stock adjustment']
        for _, adj_row in adj_df.iterrows():
            abr      = adj_row['abbreviations']
            adj_date = adj_row['date']
            adj_amt  = abs(adj_row['adjusted amount'])
            same_day_ins = report_df_sorted[
                (report_df_sorted['abbreviations'] == abr) &
                (report_df_sorted['date'] == adj_date) &
                (report_df_sorted['movement_type'] == 'Stock-in')]
            if not same_day_ins.empty:
                i = same_day_ins.index[-1]
                report_df_sorted.at[i, 'adjusted amount'] = max(0, report_df_sorted.at[i, 'adjusted amount'] - adj_amt)
                continue
            prev_ins = report_df_sorted[
                (report_df_sorted['abbreviations'] == abr) &
                (report_df_sorted['date'] < adj_date) &
                (report_df_sorted['movement_type'] == 'Stock-in')]
            if not prev_ins.empty:
                i = prev_ins.index[-1]
                report_df_sorted.at[i, 'adjusted amount'] = max(0, report_df_sorted.at[i, 'adjusted amount'] - adj_amt)
        ins_df  = report_df_sorted[report_df_sorted['movement_type'] == 'Stock-in'].groupby(
            ['date', 'abbreviations'])['adjusted amount'].sum().reset_index(name='stock_in')
        outs_df = report_df_sorted[report_df_sorted['movement_type'] == 'Invoice Issue'].groupby(
            ['date', 'abbreviations'])['adjusted amount'].sum().reset_index(name='sales')
        first_appearances = report_df_sorted.drop_duplicates(subset=['abbreviations'], keep='first')
        openings = dict(zip(first_appearances['abbreviations'], first_appearances['book quantity']))
        damages_df = damages_df[pd.notna(damages_df['quantity'])]
        wb = openpyxl.load_workbook(BytesIO(template_bytes))
        ws = wb['Sheet1']
        product_map = {}
        norm_to_abr = {}
        for r in range(1, ws.max_row + 1):
            full = ws.cell(r, 2).value
            abr  = ws.cell(r, 3).value
            if full and abr:
                product_map[str(full).strip().upper()] = abr
                norm_to_abr[normalize_name(full)] = abr
        damages_dict = {}
        for _, drow in damages_df.iterrows():
            full = str(drow['good name']).strip().upper()
            qty  = int(drow['quantity'])
            if full in product_map:
                damages_dict[product_map[full]] = qty
            else:
                norm_full = normalize_name(full)
                if norm_full in norm_to_abr:
                    damages_dict[norm_to_abr[norm_full]] = qty
        report_abr_map = {}
        for _, rrow in report_df.iterrows():
            full       = str(rrow['good name']).strip().upper()
            abr_report = rrow['abbreviations']
            if full in product_map:
                report_abr_map[abr_report] = product_map[full]
            else:
                norm_full = normalize_name(full)
                if norm_full in norm_to_abr:
                    report_abr_map[abr_report] = norm_to_abr[norm_full]
        ins_df['abbreviations']  = ins_df['abbreviations'].map(report_abr_map).fillna(ins_df['abbreviations'])
        outs_df['abbreviations'] = outs_df['abbreviations'].map(report_abr_map).fillna(outs_df['abbreviations'])
        mapped_openings = {report_abr_map.get(k, k): v for k, v in openings.items()}
        date_abr_to_row = {}
        current_date = None
        for r in range(1, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if date_val:
                current_date = excel_serial_to_date(date_val)
            abr = ws.cell(r, 3).value
            if current_date and abr:
                date_abr_to_row[(current_date, abr)] = r
        if date_abr_to_row:
            first_date = min(d[0] for d in date_abr_to_row)
            for abr, open_bal in mapped_openings.items():
                key = (first_date, abr)
                if key in date_abr_to_row:
                    ws.cell(date_abr_to_row[key], 4).value = open_bal
        damages_per_day = {}
        for abr, total_d in damages_dict.items():
            abr_ins = ins_df[ins_df['abbreviations'] == abr]
            # Only allocate damages onto days that actually have a row in
            # the template — the movement report can cover a wider date
            # range than the template (e.g. it included September stock-in
            # rows for a July-only template here), and any damages randomly
            # assigned to a day with no template row would silently vanish
            # when writing, instead of just not being drawn in the first
            # place.
            abr_ins = abr_ins[abr_ins['date'].apply(lambda d: (d.date(), abr) in date_abr_to_row)]
            if abr_ins.empty:
                continue
            days           = abr_ins['date'].dt.date.values
            stock_ins      = abr_ins['stock_in'].values
            total_stock_in = stock_ins.sum()
            if total_stock_in == 0:
                continue
            weights    = stock_ins / total_stock_in
            prod_d     = int(total_d * 3 / 4)
            pack_d     = total_d - prod_d
            prod_alloc = np.zeros(len(days), dtype=int)
            pack_alloc = np.zeros(len(days), dtype=int)
            for _ in range(prod_d):
                prod_alloc[np.random.choice(len(days), p=weights)] += 1
            for _ in range(pack_d):
                pack_alloc[np.random.choice(len(days), p=weights)] += 1
            damages_per_day[abr] = {days[i]: (prod_alloc[i], pack_alloc[i]) for i in range(len(days))}
        for _, irow in ins_df.iterrows():
            dt       = irow['date'].date()
            abr      = irow['abbreviations']
            stock_in = irow['stock_in']
            key      = (dt, abr)
            if key not in date_abr_to_row:
                continue
            row_num = date_abr_to_row[key]
            prod_d_day, pack_d_day = damages_per_day.get(abr, {}).get(dt, (0, 0))
            total_d_day = prod_d_day + pack_d_day
            ws.cell(row_num, 7).value  = stock_in + total_d_day
            ws.cell(row_num, 8).value  = prod_d_day
            ws.cell(row_num, 10).value = pack_d_day
            actual_filled = stock_in + total_d_day
            if actual_filled > 0:
                if actual_filled <= 50:    diff = random.randint(-1, 1)
                elif actual_filled <= 200: diff = random.randint(-4, 6)
                else:                      diff = random.randint(-6, 12)
                ws.cell(row_num, 6).value = max(0, actual_filled + diff)
        for _, orow in outs_df.iterrows():
            dt  = orow['date'].date()
            abr = orow['abbreviations']
            key = (dt, abr)
            if key in date_abr_to_row:
                ws.cell(date_abr_to_row[key], 13).value = orow['sales']
        output_bytes = BytesIO()
        wb.save(output_bytes)
        output_bytes.seek(0)
        return output_bytes
    except Exception as e:
        st.error(f"Error: {str(e)}")
        return None


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 2: EFRIS Invoice Enricher
# ─────────────────────────────────────────────────────────────────────────────

def _parse_pdf_bytes(pdf_bytes):
    import pdfplumber, io
    items = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                for table in (page.extract_tables() or []):
                    header_idx, col_map = None, {}
                    for i, row in enumerate(table):
                        if not row:
                            continue
                        upper = [str(c or "").upper().strip() for c in row]
                        if "ITEM" in upper and "QUANTITY" in upper:
                            header_idx = i
                            for j, h in enumerate(upper):
                                if h == "ITEM":                       col_map["item"]         = j
                                elif h == "QUANTITY":                 col_map["quantity"]     = j
                                elif "UNIT" in h and "MEASURE" in h: col_map["unit_measure"] = j
                                elif "UNIT" in h and "PRICE" in h:   col_map["unit_price"]   = j
                            continue
                        if header_idx is not None:
                            def g(key, fb, row=row):
                                ix = col_map.get(key, fb)
                                return str(row[ix] or "").strip() if ix < len(row) else ""
                            item_name = g("item", 1)
                            qty = g("quantity", 2)
                            if item_name and qty and not item_name.upper().startswith("TAX"):
                                items.append({
                                    "item": item_name,
                                    "quantity": qty,
                                    "unit_measure": g("unit_measure", 3),
                                    "unit_price": g("unit_price", 4),
                                })
                    if items:
                        return items
                text = page.extract_text() or ""
                in_d = False
                for line in text.split("\n"):
                    line = line.strip()
                    if "Section D" in line or "Goods & Services" in line:
                        in_d = True
                        continue
                    if "Section E" in line or "Tax Details" in line:
                        break
                    if not in_d:
                        continue
                    m = re.match(
                        r"\d+\.?\s+(.+?)\s+(\d[\d,]*)\s+(\S[\S\-]*)\s+([\d,]+)\s+[\d,]+",
                        line
                    )
                    if m:
                        items.append({
                            "item": m.group(1).strip(),
                            "quantity": m.group(2).strip(),
                            "unit_measure": m.group(3).strip(),
                            "unit_price": m.group(4).strip(),
                        })
    except Exception as e:
        print(f"[PDF PARSE ERROR] {e}")
    return items


def _get_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-gpu-sandbox")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-zygote")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    browser_candidates = [
        "/usr/bin/google-chrome", "/usr/bin/google-chrome-stable",
        "/usr/lib/chromium/chromium", "/usr/bin/chromium", "/usr/bin/chromium-browser",
    ]
    driver_candidates = [
        "/usr/bin/chromedriver", "/usr/local/bin/chromedriver", "/usr/lib/chromium/chromedriver",
    ]
    browser   = next((p for p in browser_candidates if os.path.exists(p)), None)
    driver_bin = next((p for p in driver_candidates if os.path.exists(p)), None)
    if not browser:
        hits = glob.glob("/usr/**/chrome", recursive=True) + glob.glob("/usr/**/chromium", recursive=True)
        browser = next((h for h in hits if os.access(h, os.X_OK)), None)
    if not driver_bin:
        hits = glob.glob("/usr/**/chromedriver", recursive=True)
        driver_bin = next((h for h in hits if os.access(h, os.X_OK)), None)
    if browser:
        options.binary_location = browser
    if driver_bin:
        return webdriver.Chrome(service=Service(executable_path=driver_bin), options=options)
    return webdriver.Chrome(options=options)


def _scrape_fdn(driver, fdn, log_fn=None):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC

    def dbg(msg):
        print(msg)
        if log_fn:
            log_fn(msg)

    # The whole EFRIS form lives inside <iframe id="ifm">, so every element
    # lookup below must happen after switching into it — the top-level
    # document has no <input>/<button> elements at all.
    items = []
    try:
        driver.get("https://efris.ura.go.ug/")
        wait = WebDriverWait(driver, 20)
        form_frame = wait.until(EC.presence_of_element_located((By.ID, "ifm")))
        driver.switch_to.frame(form_frame)
        dbg("  [1] Loaded EFRIS, entered form iframe")

        inp = wait.until(EC.presence_of_element_located((By.XPATH,
            "//input[@placeholder and ("
            "contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'fiscal')"
            " or contains(translate(@placeholder,'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'fdn')"
            ")]"
        )))
        inp.clear(); inp.send_keys(str(fdn)); time.sleep(0.4)
        dbg("  [2] FDN typed")

        # "Validate" is a <span> inside a custom button component, not a
        # real <button> element.
        validate_el = wait.until(EC.presence_of_element_located((By.XPATH,
            "//span[contains(translate(normalize-space(.),"
            "'abcdefghijklmnopqrstuvwxyz','ABCDEFGHIJKLMNOPQRSTUVWXYZ'),'VALIDATE')]"
        )))
        validate_el.click()
        dbg("  [3] Validate clicked")

        # A successful validation opens a second, dynamically-created iframe
        # at the top level (not nested inside #ifm) holding the report.
        driver.switch_to.default_content()
        report_frame = wait.until(EC.presence_of_element_located(
            (By.XPATH, "//iframe[contains(@src,'investigationValidationReport')]")))
        driver.switch_to.frame(report_frame)
        dbg("  [4] Invoice verified")

        # The invoice PDF sits at a predictable URL keyed by FDN and is
        # fetchable directly — no need to click "View Document" and sniff
        # network logs for it. Must be fetched via the page's own fetch()
        # (same-origin, carries real browser session/TLS fingerprint);
        # a plain `requests` call to this URL hangs/times out.
        pdf_url = f"https://efris.ura.go.ug/sit/investigation/validation/view?invoiceNo={fdn}"
        driver.switch_to.default_content()
        js = (
            "var cb = arguments[arguments.length-1];"
            "fetch(arguments[0], {credentials: 'include'})"
            ".then(r => r.ok ? r.arrayBuffer() : Promise.reject('HTTP ' + r.status))"
            ".then(b => cb(Array.from(new Uint8Array(b))))"
            ".catch(e => cb('ERR:' + e));"
        )
        driver.set_script_timeout(30)
        result = driver.execute_async_script(js, pdf_url)
        if isinstance(result, str):
            dbg(f"  [5] PDF fetch failed: {result}")
        else:
            pdf_bytes = bytes(result)
            dbg(f"  [5] PDF fetched ({len(pdf_bytes)} bytes)")
            items = _parse_pdf_bytes(pdf_bytes)
    except Exception as e:
        dbg(f"  [ERR] {e}")
    finally:
        try:
            driver.switch_to.default_content()
        except Exception:
            pass
    return items


def fuzzy_match(target, candidates):
    t  = target.strip().upper()
    cs = [c.strip().upper() for c in candidates]
    ms = difflib.get_close_matches(t, cs, n=1, cutoff=0.55)
    return candidates[cs.index(ms[0])] if ms else None


EFRIS_CONCURRENCY = 3  # number of invoices scraped in parallel


def run_efris_enrichment(purchases_df, log_placeholder, progress_bar):
    import threading, queue

    for col in ["Quantity", "Unit Measure", "Unit Price"]:
        if col not in purchases_df.columns:
            purchases_df[col] = None
    log_lines = []
    def log(msg):
        log_lines.append(msg)
        log_placeholder.markdown(
            '<div class="log-box">' + "<br>".join(log_lines[-80:]) + "</div>",
            unsafe_allow_html=True)
    import shutil
    for name in ["chromium", "chromium-browser", "chromedriver"]:
        p = shutil.which(name) or next(
            (c for c in [f"/usr/lib/chromium/{name}", f"/usr/bin/{name}"] if os.path.exists(c)), None)
        log(f"{'✅' if p else '❌'}  {name} → {p or 'not found'}")

    # Collect the unique FDNs that actually need scraping (first-appearance order).
    ordered_fdns, seen = [], set()
    for _, row in purchases_df.iterrows():
        fdn = str(row.get("FDN", "")).strip()
        if fdn and fdn.lower() != "nan" and fdn not in seen:
            seen.add(fdn); ordered_fdns.append(fdn)
    total_fdns = len(ordered_fdns)
    if total_fdns == 0:
        log("⚠️  No FDNs found in this file.")
        return purchases_df

    log(f"🚀 Scraping {total_fdns} unique invoice(s) with {EFRIS_CONCURRENCY} concurrent session(s)...")

    # Each worker owns one persistent browser for its whole lifetime (reused
    # across every FDN it processes, not relaunched per invoice) and pulls
    # FDNs off a shared queue until it's empty. Workers never touch the
    # Streamlit placeholders directly — only the main thread does, via
    # results_q — since Streamlit UI calls aren't safe from background
    # threads.
    work_q = queue.Queue()
    for fdn in ordered_fdns:
        work_q.put(fdn)
    results_q = queue.Queue()

    def worker():
        try:
            driver = _get_driver()
        except Exception as e:
            while True:
                try:
                    fdn = work_q.get_nowait()
                except queue.Empty:
                    break
                results_q.put((fdn, [], f"browser failed to start: {e}"))
            return
        try:
            while True:
                try:
                    fdn = work_q.get_nowait()
                except queue.Empty:
                    break
                try:
                    items = _scrape_fdn(driver, fdn, log_fn=None)
                    results_q.put((fdn, items, None))
                except Exception as e:
                    results_q.put((fdn, [], str(e)))
        finally:
            try:
                driver.quit()
            except Exception:
                pass

    threads = [threading.Thread(target=worker, daemon=True) for _ in range(EFRIS_CONCURRENCY)]
    for t in threads:
        t.start()

    fdn_cache = {}
    for i in range(total_fdns):
        fdn, items, err = results_q.get()
        fdn_cache[fdn] = items
        progress_bar.progress(min((i + 1) / total_fdns, 1.0), text=f"{i+1}/{total_fdns} invoices scraped")
        if err:
            log(f"[FDN {fdn}] ❌  {err}")
        else:
            log(f"[FDN {fdn}] ✅  {len(items)} item(s) found")
    for t in threads:
        t.join()

    log("🔗 Matching report rows to invoice items...")
    # Match against the remaining, unclaimed items for each invoice — once an
    # item is used it's removed, so two report rows with the same description
    # (e.g. an invoice listing one item twice at different quantities) each
    # get a distinct invoice line instead of both being assigned whichever
    # one matches first.
    fdn_pool = {fdn: list(items) for fdn, items in fdn_cache.items()}
    for idx, row in purchases_df.iterrows():
        fdn  = str(row.get("FDN", "")).strip()
        desc = str(row.get("Description of Goods", "")).strip()
        row_num = idx + 2
        if not fdn or fdn.lower() == "nan":
            log(f"[Row {row_num}] ⚠️  Skipped — no FDN"); continue
        invoice_items = fdn_pool.get(fdn, [])
        if not invoice_items:
            log(f"[Row {row_num}] ⚠️  No unclaimed invoice items left for '{desc}'")
            continue
        invoice_names = [i["item"] for i in invoice_items]
        matched = fuzzy_match(desc, invoice_names)
        if matched:
            hit_idx = next((i for i, it in enumerate(invoice_items)
                             if it["item"].strip().upper() == matched.strip().upper()), None)
            if hit_idx is not None:
                hit = invoice_items.pop(hit_idx)
                purchases_df.at[idx, "Quantity"]     = hit["quantity"]
                purchases_df.at[idx, "Unit Measure"] = hit["unit_measure"]
                purchases_df.at[idx, "Unit Price"]   = hit["unit_price"]
                log(f"[Row {row_num}] ✔️  '{desc}' → Qty:{hit['quantity']} Unit:{hit['unit_measure']} Price:{hit['unit_price']}")
            else:
                log(f"[Row {row_num}] ⚠️  Lookup failed: {matched}")
        else:
            log(f"[Row {row_num}] ⚠️  No match for '{desc}'")

    log("🏁 All rows processed.")
    return purchases_df


def _read_purchases_report(file_obj):
    """
    Reads a URA 'VAT Purchases Report' export (.xls or .xlsx). These exports
    have a few title/notice rows above the real header, so the header row
    is located dynamically instead of assumed to be row 0.
    """
    raw = pd.read_excel(file_obj, header=None, nrows=20)
    header_row = None
    for i in range(len(raw)):
        row_vals = [str(v).strip().upper() for v in raw.iloc[i].tolist()]
        if "FDN" in row_vals and any("DESCRIPTION OF GOODS" in v for v in row_vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            "Could not find a header row containing 'FDN' and "
            "'Description of Goods' in the first 20 rows."
        )
    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row)
    if "FDN" in df.columns:
        # FDNs exceed Excel's/int64's numeric range (the export itself warns
        # "do not convert the FDNs to number"), but not every source file
        # keeps the column as text — when it doesn't, pandas reads the huge
        # ones as genuine Python ints, and Streamlit's table preview crashes
        # trying to fit them into a fixed-width Arrow int column ("Python
        # int too large to convert to C long"). Force it to string
        # immediately so this never depends on how the source file typed
        # the column — str() on a Python int is exact, no precision lost.
        df["FDN"] = df["FDN"].apply(lambda v: str(int(v)) if isinstance(v, float) and v.is_integer() else str(v).strip())
    return df


def build_output_excel(df, sheet_name="Purchases Report",
                        highlight_cols=("Quantity", "Unit Measure", "Unit Price")):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)
        from openpyxl.styles import PatternFill, Font, Alignment
        hdr_fill = PatternFill("solid", fgColor="0078D4")
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        hi_fill  = PatternFill("solid", fgColor="FFF2CC")
        new_cols = set(highlight_cols)
        for i, c in enumerate(ws[1]):
            if c.value in new_cols:
                for row in ws.iter_rows(min_row=2, min_col=i+1, max_col=i+1):
                    for cell in row:
                        cell.fill = hi_fill
    output.seek(0)
    return output


_PACK_RE = re.compile(r'[X*]\s*(\d+)\b')


def _product_tokens(name):
    """
    Splits a product name into (size, pack_count, core_words). Used so
    stock-item matching can require pack size and case count to actually
    agree, and require the remaining brand/flavor words to overlap — a
    plain whole-string similarity ratio is unsafe here, since two genuinely
    different SKUs sharing the same boilerplate ("... 500ml x 12") often
    differ only in one short flavor word, which barely moves a
    character-level ratio (e.g. "Oner Mango 500ml x 12" vs
    "Oner Apple 500ml x 12" is ~85% similar character-for-character despite
    being a completely different product).

    The pack count ("x 12", "*12") is stripped out first so its digits
    aren't mistaken for a size; whatever number remains is treated as the
    size regardless of whether it carries an explicit unit — e.g. "Ginger
    125 x 12" and "Ginger 200 x 12" are different products even though
    neither names a unit (kg/g/etc).
    """
    upper = str(name).upper()
    pack_m = _PACK_RE.search(upper)
    pack = pack_m.group(1) if pack_m else None
    without_pack = _PACK_RE.sub(' ', upper)
    size_nums = re.findall(r'\d+(?:\.\d+)?', without_pack)
    size = size_nums[0] if size_nums else None
    core_words = [w for w in re.findall(r'[A-Z]+', without_pack) if len(w) > 1]
    return size, pack, core_words


def _match_stock_item(stock_name, candidate_descs, cutoff=0.60):
    """
    Stricter matcher for Stock List -> purchase-history lookups than
    fuzzy_match() (which is fine for its original job of picking a line item
    out of one already-known invoice, where collisions are unlikely).
    Matching one stock item against the *entire* product history is a much
    riskier search: requires detected pack size and case count to agree
    when both sides have one, and requires the remaining "core" words to
    overlap — weighted so generic/boilerplate words shared by many products
    (JAR, TINS, ASSORTED, CREAM...) count for little, and the word that
    actually distinguishes two products (a flavor name, say) counts for
    most of the score. Without that weighting, "Coconut Assorted Cream"
    would happily match "Orange Assorted Cream" on the two words they share
    while ignoring the one word that says they're different flavors.
    """
    s_size, s_pack, s_core = _product_tokens(stock_name)
    if not s_core:
        return None

    cand_tokens = [(c, _product_tokens(c)) for c in candidate_descs]
    doc_freq = {}
    for _, (_, _, c_core) in cand_tokens:
        for w in set(c_core):
            doc_freq[w] = doc_freq.get(w, 0) + 1

    def weight(w):
        return 1.0 / (1 + doc_freq.get(w, 0))

    # The word appearing in fewest candidates overall is the most
    # distinguishing one (usually the flavor/brand name) — require it to
    # actually match, so it can't be outvoted by a majority of generic
    # words that merely happen to also match (as in the Chocolate/Orange
    # Assorted Cream case above, where 3 of 4 words are shared boilerplate).
    rarest = min(s_core, key=lambda w: doc_freq.get(w, 0))

    best, best_score = None, 0.0
    for cand, (c_size, c_pack, c_core) in cand_tokens:
        if s_size and c_size and s_size != c_size:
            continue
        if s_pack and c_pack and s_pack != c_pack:
            continue
        if not c_core:
            continue
        rarest_hit = max((difflib.SequenceMatcher(None, rarest, cw).ratio() for cw in c_core), default=0)
        if rarest_hit < 0.75:
            continue
        total_w = hit_w = 0.0
        for w in s_core:
            wt = weight(w)
            total_w += wt
            if max((difflib.SequenceMatcher(None, w, cw).ratio() for cw in c_core), default=0) >= 0.75:
                hit_w += wt
        coverage = hit_w / total_w if total_w else 0.0
        if coverage < 0.6:
            continue
        whole_ratio = difflib.SequenceMatcher(None, stock_name.strip().upper(), cand.strip().upper()).ratio()
        score = (coverage + whole_ratio) / 2
        if score > best_score:
            best_score, best = score, cand
    return best if best_score >= cutoff else None


def _read_stock_list(file_obj):
    """
    Reads a URA 'Stock List' export (.xls or .xlsx). Like the purchases
    report, it has a title row above the real header.
    """
    raw = pd.read_excel(file_obj, header=None, nrows=20)
    header_row = None
    for i in range(len(raw)):
        row_vals = [str(v).strip().upper() for v in raw.iloc[i].tolist()]
        if "STOCK" in row_vals and any("GOODS/SERVICES NAME" in v for v in row_vals):
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            "Could not find a header row containing 'Goods/Services Name' and "
            "'Stock' in the first 20 rows."
        )
    file_obj.seek(0)
    return pd.read_excel(file_obj, header=header_row)


def _latest_unit_prices(enriched_df):
    """
    Builds {normalized product name: {price, date, desc}} from an enriched
    purchases report, keeping only the row with the most recent Invoice Date
    per product — i.e. the price from the latest purchase invoice for that
    good. Rows with no Unit Price (unmatched/unscraped) are ignored.
    """
    df = enriched_df.copy()
    df = df[df["Unit Price"].notna() & (df["Unit Price"].astype(str).str.strip() != "")]
    df["_date"] = pd.to_datetime(df["Invoice Date"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["_date"]).sort_values("_date")

    latest = {}
    for _, row in df.iterrows():
        try:
            price = float(str(row["Unit Price"]).replace(",", "").strip())
        except ValueError:
            continue
        norm = _norm_name(row["Description of Goods"])
        # ascending date order means later rows overwrite earlier ones,
        # so what remains is always the most recent priced purchase
        latest[norm] = {
            "price": price,
            "date": row["_date"],
            "desc": str(row["Description of Goods"]).strip(),
        }
    return latest


def value_stock(enriched_df, stock_df):
    """
    Values a stock list using the unit price from each item's latest
    purchase invoice (as enriched by the EFRIS Invoice Enricher).
    Returns (valued_df, log_lines, total_value).
    """
    log = []
    def L(msg): log.append(msg)

    L("📋 Step 1 — Building latest-price lookup from enriched purchases...")
    latest = _latest_unit_prices(enriched_df)
    L(f"  ✅ {len(latest)} distinct product(s) with a priced purchase")

    candidate_descs = [v["desc"] for v in latest.values()]

    L("🔍 Step 2 — Matching stock items to latest purchase prices...")
    out = stock_df.copy()
    for col in ["Matched Purchase Item", "Unit Price", "Latest Invoice Date", "Stock Value"]:
        out[col] = None

    matched_count = 0
    for idx, row in out.iterrows():
        name = str(row.get("Goods/Services Name", "")).strip(" .")
        if not name:
            continue
        hit_desc = _match_stock_item(name, candidate_descs) if candidate_descs else None
        if not hit_desc:
            L(f"  ✗ [NO MATCH]  '{name}'")
            continue
        info = latest.get(_norm_name(hit_desc))
        if not info:
            L(f"  ✗ [LOOKUP FAILED]  '{name}' → '{hit_desc}'")
            continue
        try:
            stock_qty = float(str(row.get("Stock", 0)).strip())
        except ValueError:
            stock_qty = 0.0
        value = round(stock_qty * info["price"], 2)
        out.at[idx, "Matched Purchase Item"]  = info["desc"]
        out.at[idx, "Unit Price"]             = info["price"]
        out.at[idx, "Latest Invoice Date"]    = info["date"].strftime("%d/%m/%Y")
        out.at[idx, "Stock Value"]            = value
        matched_count += 1
        tag = "EXACT" if _norm_name(name) == _norm_name(hit_desc) else "FUZZY"
        L(f"  ✓ [{tag:5}]  '{name}'  →  '{info['desc']}'  @ {info['price']:,}  x  {stock_qty:,}  =  {value:,.2f}")

    L(f"  ✅ {matched_count}/{len(out)} stock item(s) valued")
    total_value = out["Stock Value"].dropna().astype(float).sum()
    L(f"🏁 Total stock value: {total_value:,.2f}")
    return out, log, total_value


# ─────────────────────────────────────────────────────────────────────────────
# TOOL 3: Raw Material Movement Filler
# ─────────────────────────────────────────────────────────────────────────────

# Header / output row keywords that are never raw material or product names
_SM_HEADER_WORDS = {
    'details', 'units', 'output', 'input', 'qty', 'price', 'amount',
    'unit', 'measure', 'total', 'description', 'items', 'sub',
    'out put', 'out', 'put',
}

def _norm_name(s):
    """Uppercase + strip — used as the matching key for product names."""
    return str(s).strip().upper()


def _detect_sm_structure(df):
    """
    Dynamically locate the three key rows in ANY company's standard mix:
      product_name_row   — first row where col C holds a real product name
      abr_row            — product_name_row + 1  (abbreviations, kept for reference only)
      material_start_row — first row in col A with a real ingredient name
    Works regardless of how many header/blank rows appear before the data.
    """
    product_name_row = None
    for r in range(min(10, df.shape[0])):
        v = str(df.iloc[r, 2]).strip()
        if v in ('nan', 'NaN', ''):
            continue
        try:
            float(v); continue
        except ValueError:
            pass
        if v.lower() in _SM_HEADER_WORDS:
            continue
        product_name_row = r
        break
    if product_name_row is None:
        raise ValueError(
            "Standard mix: cannot detect product name row. "
            "Expected a product name in column C within the first 10 rows."
        )
    abr_row = product_name_row + 1

    # The "OUTPUT" row (the finished-product line, whose QTY cell is where
    # expected production belongs) explicitly marks where the raw-material
    # list starts right below it. Found by exact label match rather than by
    # skipping until the first non-header word, so its row number is known
    # precisely (needed later to resolve domestic-usage formulas that
    # reference it, e.g. "=EN8/50").
    output_row = None
    for r in range(abr_row + 1, min(abr_row + 20, df.shape[0])):
        v = str(df.iloc[r, 0]).strip()
        if v.upper() == 'OUTPUT':
            output_row = r
            break
    if output_row is None:
        raise ValueError(
            "Standard mix: cannot find the 'OUTPUT' row (the finished-product "
            "line that marks where the raw material list begins)."
        )
    material_start_row = output_row + 1
    return product_name_row, abr_row, output_row, material_start_row


def _parse_standard_mix(sm_bytes):
    """
    Parses any company's standard mix dynamically.

    Matching key = NORMALISED PRODUCT NAME (uppercase, stripped).
    ABRs are intentionally NOT used as keys because:
      - The same ABR can appear for different products (e.g. 'BB' for both
        'BANS BIG' and 'BUFFET BREAD' in some files).
      - ABR spacing is inconsistent across files ('C S' vs 'CS').
      - Product names are unambiguous and stable.

    Returns:
      products  : {norm_product_name: {'name': str, 'ratios': {mat_name: float}}}
      materials : [(row, mat_name, unit)] — row is the 1-indexed Excel row
                  (matching domestic_formulas' keys, and the row numbers
                  used inside EN<row>/EO<row> formula references), NOT the
                  0-indexed pandas row used internally while scanning.
      domestic_formulas : {excel_row_1indexed: canonicalised_formula_or_literal}
    """
    df = pd.read_excel(BytesIO(sm_bytes), header=None)
    product_name_row, abr_row, output_row, material_start_row = _detect_sm_structure(df)

    # Raw materials — col 0 = name, col 1 = unit, scanned all the way to the
    # end of the sheet rather than stopping at the first blank row. Real
    # files have formula-active-but-unnamed "reserved" rows in the middle of
    # this list (still legitimate, just unused slots), and some also list
    # more named rows further down — other finished products used as an
    # ingredient, packaging materials, labour — so a blank name can't be
    # trusted as "the list is over"; only a genuinely named row counts.
    materials = []
    for r in range(material_start_row, df.shape[0]):
        mat_name = str(df.iloc[r, 0]).strip()
        if mat_name in ('nan', 'NaN', ''):
            continue
        try:
            float(mat_name); continue
        except ValueError:
            pass
        if mat_name.lower() in _SM_HEADER_WORDS:
            continue
        mat_unit = str(df.iloc[r, 1]).strip()
        unit_clean = mat_unit if mat_unit not in ('nan', 'NaN', '') else ''
        materials.append((r + 1, mat_name, unit_clean))  # +1: pandas row -> Excel row

    if not materials:
        raise ValueError(
            "Standard mix: no raw materials found below the OUTPUT row. "
            "Check that ingredient names start in column A."
        )

    # Products — every 5 columns starting at col 2
    # Keyed by normalised product NAME to avoid all ABR ambiguity
    products = {}
    for c in range(2, df.shape[1], 5):
        name_str = str(df.iloc[product_name_row, c]).strip()
        # Skip empty, numeric, or header-keyword cells
        if name_str in ('nan', 'NaN', ''):
            continue
        try:
            float(name_str); continue
        except ValueError:
            pass
        if name_str.lower() in _SM_HEADER_WORDS:
            continue

        ratios = {}
        for row_idx, mat_name, _ in materials:
            val = df.iloc[row_idx - 1, c]  # row_idx is the 1-indexed Excel row; df.iloc is 0-indexed
            try:
                ratios[mat_name] = (
                    float(val) if str(val).strip() not in ('nan', 'NaN', '') else 0.0
                )
            except (ValueError, TypeError):
                ratios[mat_name] = 0.0

        norm = _norm_name(name_str)
        # If two columns share the same normalised name, last one wins (edge case)
        products[norm] = {'name': name_str, 'ratios': ratios}

    if not products:
        raise ValueError(
            "Standard mix: no valid products found. "
            "Check that product names appear in the first detected row."
        )

    domestic_formulas = _parse_domestic_usage_formulas(sm_bytes, output_row, df.shape[0])

    return products, materials, domestic_formulas


def _parse_domestic_usage_formulas(sm_bytes, output_row, sheet_row_count):
    """
    Locates the 'SUMMARY QTY CONSUMED IN PRODUCTION' and 'DOMESTIC USAGE'
    columns by header text (their position shifts depending on how many
    product blocks a given bakery's sheet has, so they can't be hardcoded),
    and returns each row's DOMESTIC USAGE formula from the OUTPUT row
    downward — with references to those two columns rewritten to canonical
    EN<row>/EO<row> tokens so the evaluator never needs to know the real
    column letters.

    Returns {excel_row_1indexed: formula_text} — formula_text has no
    leading '='. Rows with a plain hardcoded number get that number as a
    string; rows with nothing in the DOMESTIC USAGE column are omitted
    (treated as 0 by the evaluator). Returns {} if this standard mix has no
    domestic-usage section at all.
    """
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(BytesIO(sm_bytes), data_only=False)
    ws = wb.active

    summary_col = domestic_col = None
    for r in range(1, min(10, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            vu = v.strip().upper()
            if vu == 'SUMMARY QTY CONSUMED IN PRODUCTION':
                summary_col = c
            elif vu == 'DOMESTIC USAGE':
                domestic_col = c
        if summary_col and domestic_col:
            break

    if domestic_col is None:
        return {}  # this bakery's standard mix has no domestic-usage section

    domestic_letter = get_column_letter(domestic_col)
    ref_letters = [domestic_letter] + ([get_column_letter(summary_col)] if summary_col else [])
    ref_re = re.compile(r'\$?(' + '|'.join(re.escape(l) for l in ref_letters) + r')\$?(\d+)\b')

    def canon(m):
        col, row = m.group(1), m.group(2)
        return (f'EO{row}' if col == domestic_letter else f'EN{row}')

    formulas = {}
    start_row = output_row + 1  # openpyxl is 1-indexed; output_row (pandas 0-indexed) + 1 == its own Excel row
    end_row = min(sheet_row_count, ws.max_row)
    for r in range(start_row, end_row + 1):
        v = ws.cell(r, domestic_col).value
        if isinstance(v, str) and v.startswith('='):
            formulas[r] = ref_re.sub(canon, v[1:])
        elif isinstance(v, (int, float)):
            formulas[r] = str(v)
    return formulas


def _eval_domestic_formula(row, domestic_formulas, production_by_row, memo, visiting=None):
    """
    Evaluates a canonicalised DOMESTIC USAGE formula for one standard-mix
    row, resolving:
      EN<row> -> that row's already-computed production consumption
                 for the day currently being processed
      EO<row> -> that row's own domestic-usage formula, evaluated recursively
    Only basic arithmetic (+ - * / and parentheses) plus these two reference
    kinds are supported — a formula using anything else (a spreadsheet
    function, say) evaluates to 0.0 rather than guessing at it.
    """
    if row in memo:
        return memo[row]
    formula = domestic_formulas.get(row)
    if formula is None:
        return 0.0
    if visiting is None:
        visiting = set()
    if row in visiting:
        return 0.0  # circular reference guard
    visiting.add(row)

    def replace_ref(m):
        col, ref_row = m.group(1), int(m.group(2))
        if col == 'EN':
            return repr(production_by_row.get(ref_row, 0.0))
        return repr(_eval_domestic_formula(ref_row, domestic_formulas, production_by_row, memo, visiting))

    substituted = re.sub(r'(EN|EO)(\d+)', replace_ref, formula)
    visiting.discard(row)

    if not re.fullmatch(r'[0-9.\+\-\*/() \t]*', substituted):
        memo[row] = 0.0
        return 0.0
    try:
        result = float(eval(substituted, {"__builtins__": {}}, {}))
    except Exception:
        result = 0.0
    memo[row] = result
    return result


def _parse_finished_movement(fm_bytes):
    """
    Reads col B (product name) and col F (EXPECTED) for every row.
    expected_map  keyed by (date, NORM_PRODUCT_NAME)
    fm_products   set of original product name strings (for display in errors)

    Daily blocks carry a real date in col A only on their first row, with
    every following row of that block left blank (col A keeps meaning
    "same date as above" until a new date appears) — so a blank date can't
    just be skipped, it has to inherit the last real date seen. But a
    trailing monthly-TOTAL block breaks that assumption: it starts with a
    text label ("TOTAL") in col A, not a date, and everything below it is
    also blank — so naively continuing to inherit "the last real date"
    silently folds the entire month's totals into that last day. Once col A
    holds something that isn't a date and isn't blank, everything from
    there on is treated as outside the daily data (parsing stops).
    """
    wb = openpyxl.load_workbook(BytesIO(fm_bytes), data_only=True)
    ws = wb.active
    expected_map = {}
    fm_products  = set()
    current_date = None

    for r in range(1, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        name_v = ws.cell(r, 2).value   # col B — product full name
        exp_v  = ws.cell(r, 6).value   # col F — EXPECTED

        if isinstance(date_v, datetime):
            current_date = date_v.date()
        elif current_date is not None and date_v is not None and str(date_v).strip():
            break  # a non-blank, non-date label (e.g. "TOTAL") ends the daily data —
            # only once we're past the header, i.e. a real date has already been seen

        if not name_v or not str(name_v).strip():
            continue
        name_str = str(name_v).strip()

        # Skip obvious header rows
        if name_str.upper() in ('PRODUCTS', 'DETAILS', 'DATE', 'DESCRIPTION'):
            continue

        fm_products.add(name_str)
        if current_date is None:
            continue

        try:
            qty = float(exp_v) if exp_v not in (None, '') else 0.0
        except (ValueError, TypeError):
            qty = 0.0

        key = (current_date, _norm_name(name_str))
        expected_map[key] = expected_map.get(key, 0.0) + qty

    return expected_map, fm_products


def _match_product_name(fm_name, sm_norm_names):
    """
    Match a single finished-movement product name to a standard-mix product
    name. Returns (matched_norm_name_or_None, tag_string).

    A plain whole-string fuzzy ratio is unsafe here: "TOSS BREAD HALF" and
    "TOSS BREAD BIG" share 10 of ~14 characters ("TOSS BREAD"), so a
    character-level ratio rates them as very similar even though "Half"
    vs "Big" makes them different products with different recipes. Instead,
    every word that's rare across the candidate list (appears in at most
    one of them — usually a size/variant qualifier, not generic words like
    "BREAD" or "CAKES") must actually match before a candidate is even
    considered — gating on just the single rarest word isn't enough when
    two words are tied for rarest (as TOSS and HALF are here), since the
    common word can then drag a weighted-coverage score just over the line.
    If nothing clears the bar, it's reported as NO MATCH rather than
    guessed at — the caller already surfaces unmatched products so they can
    be added to the standard mix, which is safer than silently using the
    wrong recipe.
    """
    t = _norm_name(fm_name)
    if t in sm_norm_names:
        return t, 'EXACT'

    def tokens(s):
        return [w for w in re.findall(r'[A-Z0-9]+', s) if len(w) > 1]

    t_words = tokens(t)
    if not t_words:
        return None, 'NO MATCH'

    cand_tokens = [(name, tokens(name)) for name in sm_norm_names]
    doc_freq = {}
    for _, c_words in cand_tokens:
        for w in set(c_words):
            doc_freq[w] = doc_freq.get(w, 0) + 1
    critical = [w for w in t_words if doc_freq.get(w, 0) <= 1]

    def word_hit(w, c_words):
        return max((difflib.SequenceMatcher(None, w, cw).ratio() for cw in c_words), default=0) >= 0.75

    best, best_score = None, 0.0
    for cand, c_words in cand_tokens:
        if not c_words:
            continue
        if critical and any(not word_hit(w, c_words) for w in critical):
            continue
        total_w = hit_w = 0.0
        for w in t_words:
            wt = 1.0 / (1 + doc_freq.get(w, 0))
            total_w += wt
            if word_hit(w, c_words):
                hit_w += wt
        coverage = hit_w / total_w if total_w else 0.0
        if coverage < 0.7:
            continue
        whole_ratio = difflib.SequenceMatcher(None, t, cand).ratio()
        score = (coverage + whole_ratio) / 2
        if score > best_score:
            best_score, best = score, cand

    if best and best_score >= 0.60:
        return best, 'FUZZY'
    return None, 'NO MATCH'


def _parse_rm_template(rm_bytes):
    """
    Reads the user's blank raw material template.
    Detects block structure (materials per day) dynamically — works for any
    number of materials (1 product kombucha up to 50+ product bakery).
    """
    wb = openpyxl.load_workbook(BytesIO(rm_bytes))
    ws = wb.active

    # First date row = start of day-1 block
    first_data_row = None
    for r in range(1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, datetime):
            first_data_row = r
            break
    if first_data_row is None:
        raise ValueError("Raw material template: no date found in column A.")

    # Second date row → block_size
    second_date_row = None
    for r in range(first_data_row + 1, ws.max_row + 1):
        if isinstance(ws.cell(r, 1).value, datetime):
            second_date_row = r
            break

    if second_date_row:
        block_size = second_date_row - first_data_row
    else:
        # Single-day template: count until first blank in col B
        block_size = 0
        for r in range(first_data_row, ws.max_row + 1):
            if ws.cell(r, 2).value is None and r > first_data_row:
                block_size = r - first_data_row + 1
                break
        if block_size == 0:
            block_size = ws.max_row - first_data_row + 1

    n_materials = block_size - 1   # last row in every block is the blank separator

    # Read material names + units from day-1 block (col B = name, col C = unit)
    tpl_materials = []
    for i in range(n_materials):
        r   = first_data_row + i
        mat = ws.cell(r, 2).value
        unt = ws.cell(r, 3).value
        if mat and str(mat).strip():
            tpl_materials.append((
                i,
                str(mat).strip(),
                str(unt).strip() if unt else ''
            ))

    # Map every date in the template to its block-start row
    date_to_block_start = {}
    for r in range(first_data_row, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, datetime):
            date_to_block_start[v.date()] = r

    return wb, ws, n_materials, tpl_materials, date_to_block_start


def _smart_match_materials(tpl_materials, sm_materials):
    """
    Match raw material names in the template → raw material names in the
    standard mix.  Uses exact then fuzzy (cutoff 0.65).
    Returns match_map {tpl_name: sm_name_or_None}, unmatched list.
    """
    sm_names  = [m[1] for m in sm_materials]
    sm_upper  = [m.upper().strip() for m in sm_names]
    match_map = {}
    unmatched = []

    for _, tpl_name, _ in tpl_materials:
        t = tpl_name.upper().strip()
        matched = None

        # 1. Exact
        if t in sm_upper:
            matched = sm_names[sm_upper.index(t)]

        # 2. Fuzzy — returns best match; no substring shortcut to avoid wrong greedy hits
        if not matched:
            hits = difflib.get_close_matches(t, sm_upper, n=1, cutoff=0.65)
            if hits:
                matched = sm_names[sm_upper.index(hits[0])]

        match_map[tpl_name] = matched
        if not matched:
            unmatched.append(tpl_name)

    return match_map, unmatched


def _calculate_and_fill(wb, ws, tpl_materials, date_to_block_start,
                         match_map, products, expected_map, fm_name_to_sm_norm,
                         materials, domestic_formulas):
    """
    For each day:
      1. Production consumption is computed for EVERY row in the standard
         mix's material list, not just the ones the template has a slot
         for — domestic-usage formulas can reference *other* materials'
         production totals (this bakery's domestic oil/fat/salt usage are
         all defined as multiples of flour's production total), so all of
         them need to be available before any domestic figure can be
         evaluated.
      2. For each raw material the template actually has a slot for:
           col 9 = production consumption + domestic usage
         where domestic usage is evaluated from the standard mix's own
         DOMESTIC USAGE column formulas (see _eval_domestic_formula) rather
         than invented — this bakery already defines exactly what its
         domestic usage should be, the same way it defines material ratios.

    products      keyed by norm_product_name
    expected_map  keyed by (date, norm_product_name_from_fm)
    fm_name_to_sm_norm maps norm FM product name → norm SM product name
                        (handles fuzzy name differences between files)
    materials     [(row, mat_name, unit)] — every named row below OUTPUT
    Only col 9 is written; all other formulas in the template are preserved.
    """
    name_to_row = {mat_name: row for row, mat_name, _ in materials}

    filled = 0
    for date_obj, block_start in sorted(date_to_block_start.items()):
        production_by_row = {}
        for row_idx, mat_name, _ in materials:
            total = 0.0
            for norm_fm, exp in expected_map.items():
                if norm_fm[0] != date_obj or exp == 0.0:
                    continue
                sm_norm = fm_name_to_sm_norm.get(norm_fm[1])
                if sm_norm is None:
                    continue
                prod = products.get(sm_norm)
                if prod is None:
                    continue
                ratio = prod['ratios'].get(mat_name, 0.0)
                total += ratio * exp
            production_by_row[row_idx] = total

        memo = {}
        for offset, tpl_name, _ in tpl_materials:
            sm_mat_name = match_map.get(tpl_name)
            if sm_mat_name is None:
                continue
            row_idx = name_to_row.get(sm_mat_name)
            if row_idx is None:
                continue

            production = production_by_row.get(row_idx, 0.0)
            domestic = _eval_domestic_formula(row_idx, domestic_formulas, production_by_row, memo)
            total = production + domestic

            if total > 0:
                ws.cell(block_start + offset, 9).value = round(total, 4)
                filled += 1

    return wb, filled


def process_raw_material_movement(sm_file, fm_file, rm_template_file, output_name):
    """Master function called from Streamlit UI."""
    log = []
    def L(msg): log.append(msg)

    try:
        # ── Step 1: Standard Mix ─────────────────────────────────────────────
        L("📋 Step 1 — Parsing standard mix...")
        sm_bytes  = sm_file.read()
        products, sm_materials, domestic_formulas = _parse_standard_mix(sm_bytes)
        L(f"  ✅ {len(products)} finished product(s) | {len(sm_materials)} named line(s) in "
          f"standard mix (ingredients, and possibly packaging/cross-product/labour if used)")
        if domestic_formulas:
            L(f"  ✅ Domestic usage section found — {len(domestic_formulas)} row(s) with a formula")
        else:
            L("  ℹ️  No domestic usage section found in this standard mix — domestic usage will be 0")
        L("  📦 Products in standard mix:")
        for norm, p in products.items():
            L(f"       • {p['name']}")

        # ── Step 2: Finished Movement ────────────────────────────────────────
        L("📋 Step 2 — Parsing finished movement (EXPECTED column)...")
        fm_bytes  = fm_file.read()
        expected_map, fm_products = _parse_finished_movement(fm_bytes)
        non_zero  = sum(1 for v in expected_map.values() if v > 0)
        L(f"  ✅ {len(fm_products)} product name(s) | "
          f"{len(expected_map)} date-product pairs ({non_zero} with data)")

        # ── Step 3: Match FM product names → SM product names ────────────────
        L("🔍 Step 3 — Matching finished movement products to standard mix...")
        sm_norm_list = list(products.keys())   # normalised SM product names
        fm_name_to_sm_norm = {}
        missing = []
        fuzzy_warnings = []

        for fm_name in sorted(fm_products):
            matched_norm, tag = _match_product_name(fm_name, sm_norm_list)
            if matched_norm:
                fm_name_to_sm_norm[_norm_name(fm_name)] = matched_norm
                sm_display = products[matched_norm]['name']
                L(f"       ✓ [{tag:8}]  '{fm_name}'  →  '{sm_display}'")
                if tag == 'FUZZY':
                    fuzzy_warnings.append((fm_name, sm_display))
            else:
                missing.append(fm_name)
                L(f"       ✗ [NO MATCH]  '{fm_name}'  →  NOT FOUND in standard mix")

        if missing:
            L(f"  ❌ {len(missing)} product(s) from the finished movement could not be "
              f"matched to any product in the standard mix:")
            for m in missing:
                L(f"       • '{m}'")
            L("  ⚠️  Add these products to the standard mix before processing.")
            return None, log

        if fuzzy_warnings:
            L(f"  ⚠️  {len(fuzzy_warnings)} product(s) matched by fuzzy name — "
              f"please verify these are correct:")
            for fm_n, sm_n in fuzzy_warnings:
                L(f"       '{fm_n}'  →  '{sm_n}'")

        L(f"  ✅ All {len(fm_products)} product(s) matched successfully")

        # ── Step 4: Raw Material Template ────────────────────────────────────
        L("📋 Step 4 — Reading raw material template structure...")
        rm_bytes  = rm_template_file.read()
        wb, ws, n_materials, tpl_materials, date_to_block_start = _parse_rm_template(rm_bytes)
        L(f"  ✅ {len(date_to_block_start)} day(s) | {n_materials} material row(s) per day")
        L("  📄 Template materials:")
        for _, name, unit in tpl_materials:
            L(f"       • {name} ({unit})")

        # ── Step 5: Match template material names → SM material names ────────
        L("🔗 Step 5 — Matching template materials to standard mix...")
        match_map, unmatched = _smart_match_materials(tpl_materials, sm_materials)
        matched_n = sum(1 for v in match_map.values() if v)
        L(f"  ✅ {matched_n}/{len(tpl_materials)} raw material(s) matched")
        for tpl_n, sm_n in match_map.items():
            if sm_n:
                tag = "EXACT" if tpl_n.upper().strip() == sm_n.upper().strip() else "FUZZY"
                L(f"       ✓ [{tag}]  '{tpl_n}'  →  '{sm_n}'")
            else:
                L(f"       ✗ [NO MATCH]  '{tpl_n}'  →  will be left blank in output")
        if unmatched:
            L(f"  ⚠️  {len(unmatched)} raw material(s) unmatched — "
              f"verify spelling between template and standard mix")

        # ── Step 6: Calculate and fill ───────────────────────────────────────
        L("⚙️  Step 6 — Calculating STOCK ISSUED TO PRODUCTION (production + domestic usage)...")
        wb_out, filled = _calculate_and_fill(
            wb, ws, tpl_materials, date_to_block_start,
            match_map, products, expected_map, fm_name_to_sm_norm,
            sm_materials, domestic_formulas
        )
        L(f"  ✅ {filled} cell(s) written to STOCK ISSUED TO PRODUCTION (col I)")

        out = BytesIO()
        wb_out.save(out)
        out.seek(0)
        L(f"✅ Complete — ready to download: {output_name}")
        return out, log

    except Exception as e:
        import traceback
        L(f"❌ Error: {str(e)}")
        L(traceback.format_exc())
        return None, log


# ─────────────────────────────────────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────────────────────────────────────

st.sidebar.title("Navigation")
tool = st.sidebar.selectbox("Select Tool", [
    "Excel Stock Movement Filler",
    "EFRIS Invoice Enricher",
    "Raw Material Movement Filler",
    "Audit Compliance Checker (Coming Soon)",
    "Financial Report Generator (Coming Soon)",
    "Sales Dashboard (Coming Soon)",
])

st.title("Automation Hub")
st.markdown("Your professional platform for automating tasks.")


# ── TOOL 1 UI ────────────────────────────────────────────────────────────────
if tool == "Excel Stock Movement Filler":
    st.header("Excel Stock Movement Filler")
    output_name   = st.text_input("Output Filename", value="filled_template")
    output_name   = output_name.removesuffix(".xlsx").strip() + ".xlsx"
    template_file = st.file_uploader("Upload Template (.xlsx)", type="xlsx")
    report_file   = st.file_uploader("Upload Movement Report (.xlsx)", type="xlsx")
    damages_file  = st.file_uploader("Upload Damages (.xlsx)", type="xlsx")
    if st.button("Process Files"):
        if template_file and report_file and damages_file:
            with st.spinner("Processing..."):
                out = process_excel(template_file, report_file, damages_file, output_name)
                if out:
                    st.success("Done!")
                    st.download_button("Download", data=out, file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        else:
            st.warning("Upload all 3 files.")


# ── TOOL 2 UI ────────────────────────────────────────────────────────────────
elif tool == "EFRIS Invoice Enricher":
    st.header("EFRIS Invoice Enricher")
    st.markdown("""
    Upload your **Purchases Report** (.xlsx or .xls — the URA "VAT Purchases Report"
    export works directly) with columns **FDN** and **Description of Goods**.
    The tool opens each invoice on EFRIS, reads the PDF, and fills in
    **Quantity**, **Unit Measure**, and **Unit Price** for every row.
    > Duplicate FDNs are only fetched once.
    """)
    col1, col2 = st.columns([2, 1])
    with col1:
        purchases_file = st.file_uploader("Upload Purchases Report (.xlsx or .xls)", type=["xlsx", "xls"], key="ef_up")
    with col2:
        out_name = st.text_input("Output Filename", value="enriched_purchases", key="ef_out")
        out_name = out_name.removesuffix(".xlsx").strip() + ".xlsx"
    if purchases_file:
        try:
            prev = _read_purchases_report(purchases_file)
            purchases_file.seek(0)
            st.markdown("**Preview:**")
            st.dataframe(prev.head(5), use_container_width=True)
            missing = {"FDN", "Description of Goods"} - set(prev.columns)
            if missing:
                st.error(f"Missing columns: {missing}")
                purchases_file = None
        except Exception as e:
            st.error(str(e)); purchases_file = None
    if st.button("🚀 Start Enrichment", disabled=(purchases_file is None), key="ef_run"):
        st.markdown("---")
        prog   = st.progress(0, text="Starting...")
        log_ph = st.empty()
        try:
            df       = _read_purchases_report(purchases_file)
            enriched = run_efris_enrichment(df, log_ph, prog)
            prog.progress(1.0, text="✅ Done!")
            st.session_state["efris_enriched_df"] = enriched
            st.session_state["efris_out_name"]    = out_name
        except Exception as e:
            st.error(f"Error: {e}")

    # Kept outside the button block (gated on session_state instead) so the
    # download button and the stock-valuation prompt below both survive
    # later reruns — e.g. uploading the Stock List file — instead of
    # disappearing the moment a different widget is interacted with.
    if "efris_enriched_df" in st.session_state:
        enriched = st.session_state["efris_enriched_df"]
        st.success("Complete!")
        st.download_button("⬇️ Download Enriched Excel",
            data=build_output_excel(enriched),
            file_name=st.session_state.get("efris_out_name", out_name),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="ef_download")
        filled = enriched["Quantity"].notna().sum()
        st.info(f"📊 {filled} / {len(enriched)} rows enriched.")

    # Value Stock is intentionally NOT gated on having just run an
    # enrichment in this session — if you already have a previously
    # downloaded enriched Purchases Report, you can upload it directly and
    # skip re-running EFRIS enrichment entirely.
    st.markdown("---")
    st.subheader("📦 Value Stock")
    st.markdown("""
    Value your closing stock using the **unit price from each item's latest
    purchase invoice**.
    """)

    has_session_data = "efris_enriched_df" in st.session_state
    if has_session_data:
        source_choice = st.radio(
            "Enriched purchases data",
            ["Use the result from the enrichment above", "Upload an enriched Purchases Report instead"],
            key="sv_source")
    else:
        source_choice = "Upload an enriched Purchases Report instead"
        st.caption("No enrichment run yet this session — upload a previously-downloaded "
                   "enriched Purchases Report (the file with Quantity/Unit Measure/Unit "
                   "Price already filled in) to continue.")

    enriched_for_valuation = None
    if source_choice == "Use the result from the enrichment above":
        enriched_for_valuation = st.session_state["efris_enriched_df"]
    else:
        enriched_upload = st.file_uploader(
            "Upload enriched Purchases Report (.xlsx or .xls)",
            type=["xlsx", "xls"], key="sv_enriched_up")
        if enriched_upload:
            try:
                enriched_for_valuation = pd.read_excel(enriched_upload)
                missing = {"Description of Goods", "Unit Price", "Invoice Date"} - set(enriched_for_valuation.columns)
                if missing:
                    st.error(f"Missing columns in enriched file: {missing}")
                    enriched_for_valuation = None
            except Exception as e:
                st.error(str(e))

    stock_file = st.file_uploader("Upload Stock List (.xlsx or .xls)", type=["xlsx", "xls"], key="sv_up")
    if st.button("💰 Value Stock", disabled=(stock_file is None or enriched_for_valuation is None), key="sv_run"):
        try:
            stock_df = _read_stock_list(stock_file)
            valued_df, val_log, total_value = value_stock(enriched_for_valuation, stock_df)
            st.markdown('<div class="log-box">' + "<br>".join(val_log) + "</div>", unsafe_allow_html=True)
            st.success(f"✅ Stock valued — total value: UGX {total_value:,.2f}")
            st.download_button("⬇️ Download Valued Stock List",
                data=build_output_excel(
                    valued_df, sheet_name="Stock Valuation",
                    highlight_cols=("Matched Purchase Item", "Unit Price",
                                     "Latest Invoice Date", "Stock Value")),
                file_name="stock_valuation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="sv_download")
        except Exception as e:
            st.error(f"Error: {e}")


# ── TOOL 3 UI ────────────────────────────────────────────────────────────────
elif tool == "Raw Material Movement Filler":
    st.header("Raw Material Movement Filler")
    st.markdown("""
    Calculates **STOCK ISSUED TO PRODUCTION** for every raw material, every day —
    derived from the **EXPECTED** quantities in the finished goods movement and your
    company's **standard mix** ratios.

    **How it works:**
    - For each day and each raw material:
      **Issued = Σ (input/unit ratio × EXPECTED quantity)** across all finished products
    - The result is written into the *STOCK ISSUED TO PRODUCTION* column of your
      raw material template. All existing formulas (Bal b/f carry-forwards, totals, etc.)
      are preserved.

    > ⚠️  All products in the finished movement **must** be present in the standard mix.
    > The tool will list any missing ones and stop before making changes.
    """)

    st.markdown("---")
    c1, c2, c3 = st.columns(3)
    with c1:
        sm_file = st.file_uploader(
            "① Standard Mix (.xlsx)",
            type="xlsx", key="rm_sm",
            help="Your company's standard mix with products and raw material ratios."
        )
    with c2:
        fm_file = st.file_uploader(
            "② Finished Goods Movement (.xlsx)",
            type="xlsx", key="rm_fm",
            help="Output of Tool 1 — must contain the EXPECTED column (col F)."
        )
    with c3:
        rm_file = st.file_uploader(
            "③ Raw Material Template (.xlsx)",
            type="xlsx", key="rm_tpl",
            help="Your blank raw material movement template (company-specific)."
        )

    out_name_rm = st.text_input("Output Filename", value="raw_material_filled", key="rm_out")
    out_name_rm = out_name_rm.removesuffix(".xlsx").strip() + ".xlsx"

    # Show previews of each uploaded file
    if sm_file:
        try:
            sm_prev = pd.read_excel(sm_file, header=None, nrows=4)
            sm_file.seek(0)
            with st.expander("Preview: Standard Mix (first 4 rows)"):
                st.dataframe(sm_prev, use_container_width=True)
        except Exception:
            pass

    if fm_file:
        try:
            fm_prev = pd.read_excel(fm_file, header=None, nrows=8)
            fm_file.seek(0)
            with st.expander("Preview: Finished Goods Movement (first 8 rows)"):
                st.dataframe(fm_prev, use_container_width=True)
        except Exception:
            pass

    if rm_file:
        try:
            rm_prev = pd.read_excel(rm_file, header=None, nrows=8)
            rm_file.seek(0)
            with st.expander("Preview: Raw Material Template (first 8 rows)"):
                st.dataframe(rm_prev, use_container_width=True)
        except Exception:
            pass

    all_uploaded = sm_file and fm_file and rm_file
    if not all_uploaded:
        st.info("Upload all three files above to proceed.")

    if st.button("⚙️  Generate Raw Material Movement", disabled=not all_uploaded, key="rm_run"):
        st.markdown("---")
        log_ph = st.empty()
        with st.spinner("Processing..."):
            out_bytes, log_lines = process_raw_material_movement(
                sm_file, fm_file, rm_file, out_name_rm
            )
        # Show log
        log_ph.markdown(
            '<div class="log-box">' + "<br>".join(log_lines) + "</div>",
            unsafe_allow_html=True
        )
        if out_bytes:
            st.success("✅ Processing complete!")
            st.download_button(
                "⬇️ Download Filled Raw Material Movement",
                data=out_bytes,
                file_name=out_name_rm,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Processing failed. See the log above for details.")


# ── COMING SOON ──────────────────────────────────────────────────────────────
elif "Coming Soon" in tool:
    st.info("Feature coming soon.")

st.sidebar.markdown("---")
st.sidebar.info("Powered by Streamlit on Railway.")
